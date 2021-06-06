"""
Data project

"""
from openpyxl import load_workbook, Workbook


class DataAccess:
    """Class has all the methods to access data in project.xlsx
    Has methods to view marks, hobbies, travel history, language
    expertise and domain
    """
    def __init__(self):
        self.work_book = load_workbook(filename='../xlsx_file/project.xlsx')

    def marks(self, ps_number):
        """Returns marks corresponding to entered ps number
        :parameter  ps_number
        :returns 1
        :rtype int
        """
        work_obj = self.work_book["Marks"]
        list_0 = []
        list_1 = []
        for row in work_obj.rows:
            for val in row:
                if row[0].value == "PS_Number":
                    list_0.append(val.value)
                if row[0].value == ps_number:
                    list_1.append(val.value)
        print(list_0)
        print(list_1)
        res_wb = Workbook()
        new_wb = res_wb.active
        new_wb.title = "result"
        new_wb.append(list_0)
        new_wb.append(list_1)
        res_wb.save("result.xlsx")
        print("Data written successfully into result.xlsx")
        return 1

    def hobbies(self, ps_number):
        """Returns hobbies corresponding to entered ps number
        :parameter ps_number
        :returns 2
        :rtype int
        """
        work_obj = self.work_book["Hobbies"]
        list_0 = []
        list_1 = []
        for row in work_obj.rows:
            for val in row:
                if row[0].value == "PS_Number":
                    list_0.append(val.value)
                if row[0].value == ps_number:
                    list_1.append(val.value)
        print(list_0)
        print(list_1)
        res_wb = Workbook()
        new_wb = res_wb.active
        new_wb.title = "result"
        new_wb.append(list_0)
        new_wb.append(list_1)
        res_wb.save("result.xlsx")
        print("Data written successfully into result.xlsx")
        return 2

    def travel(self, ps_number):
        """Returns travel history corresponding to entered ps number
        :parameter ps_number
        :returns 3
        :rtype int
        """
        work_obj = self.work_book["Travel"]
        list_0 = []
        list_1 = []
        for row in work_obj.rows:
            for val in row:
                if row[0].value == "PS_Number":
                    list_0.append(val.value)
                if row[0].value == ps_number:
                    list_1.append(val.value)
        print(list_0)
        print(list_1)
        res_wb = Workbook()
        new_wb = res_wb.active
        new_wb.title = "result"
        new_wb.append(list_0)
        new_wb.append(list_1)
        res_wb.save("result.xlsx")
        print("Data written successfully into result.xlsx")
        return 3

    def languages(self, ps_number):
        """Returns expertise in languages corresponding to entered ps number
        :parameter ps_number
        :returns 4
        :rtype int
        """
        work_obj = self.work_book["Languages"]
        list_0 = []
        list_1 = []
        for row in work_obj.rows:
            for val in row:
                if row[0].value == "PS_Number":
                    list_0.append(val.value)
                if row[0].value == ps_number:
                    list_1.append(val.value)
        print(list_0)
        print(list_1)
        res_wb = Workbook()
        new_wb = res_wb.active
        new_wb.title = "result"
        new_wb.append(list_0)
        new_wb.append(list_1)
        res_wb.save("result.xlsx")
        print("Data written successfully into result.xlsx")
        return 4

    def domain(self, ps_number):
        """Returns domain info corresponding to entered ps number
        :parameter ps_number
        :returns 5
        :rtype int
        """
        work_obj = self.work_book["Domain"]
        list_0 = []
        list_1 = []
        for row in work_obj.rows:
            for val in row:
                if row[0].value == "PS_Number":
                    list_0.append(val.value)
                if row[0].value == ps_number:
                    list_1.append(val.value)
        print(list_0)
        print(list_1)
        res_wb = Workbook()
        new_wb = res_wb.active
        new_wb.title = "result"
        new_wb.append(list_0)
        new_wb.append(list_1)
        res_wb.save("result.xlsx")
        print("Data written successfully into result.xlsx")
        return 5


def main(ps_num, value):
    """Checks whether the details entered are correct or not"""
    if ps_num < 1200:
        print("Enter a ps number greater than or equal to 1200")
        return 1
    elif ps_num > 1214:
        print("Enter a ps number less than 1215")
        return 2
    elif not isinstance(value, int):
        print("Enter an integer")
        return 3
    elif value <= 0:
        print("Enter a value greater than 0")
        return 4
    elif value > 5:
        print("Enter a value less than 6")
        return 5
    else:
        if value == 1:
            obj.marks(ps_num)
        elif value == 2:
            obj.hobbies(ps_num)
        elif value == 3:
            obj.travel(ps_num)
        elif value == 4:
            obj.languages(ps_num)
        else:
            obj.domain(ps_num)
        return 6


obj = DataAccess()


def entry():
    """Checking for correct datatype from user entry"""
    try:
        ps1 = int(input("Enter a four digit PS_Number "
                        "ranging from 1200-1214"))
        value1 = int(input("Enter 1 for Marks Sheet"
                           "\nEnter 2 for Hobbies"
                           "\nEnter 3 for Travel history"
                           "\nEnter 4 for Language expertise"
                           "\nEnter 5 for Domain specific information"))
        return ps1, value1
    except ValueError:
        print("Enter an integer")
        entry()
        return 0


if __name__ == "__main__":
    psf, value_final = entry()
    main(psf, value_final)
